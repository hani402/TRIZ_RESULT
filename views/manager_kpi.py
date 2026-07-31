import streamlit as st
import numpy as np
from openpyxl import Workbook

from data_utils import (
    get_managers, get_active_months, build_manager_actuals,
    load_manager_kpi_targets, MONTH_ORDER, QUARTER_MAP,
)
import excel_export as xx


def _fmt_money(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    return f"{int(round(v)):,}원"


def _fmt_pct(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    return f"{v * 100:,.2f}%"


def _safe_div(a, b):
    if b is None or (isinstance(b, float) and np.isnan(b)) or b == 0:
        return np.nan
    return a / b


def render(df):
    st.title("🧑‍💼 매니저별")

    if df is None:
        st.info("왼쪽 사이드바에서 ALL데이터 엑셀 파일을 업로드하면 결과가 표시됩니다.")
        return

    managers = get_managers(df)
    active_months = get_active_months(df)

    if not managers:
        st.warning("담당자('비고1') 정보가 있는 데이터가 없어요.")
        return

    actuals = build_manager_actuals(df, managers)  # index=(담당자,지표), columns=합계+월

    # ---- KPI 목표 (담당자별 고정 KPI 백데이터, 읽기 전용) ----
    kpi_targets = load_manager_kpi_targets()
    if not kpi_targets:
        st.warning("KPI 백데이터 파일(`data/kpi_manager_targets.xlsx`)을 찾지 못했어요. KPI 값 없이 실적만 표시할게요.")

    def _kpi_value(manager, metric, col):
        m = kpi_targets.get(manager)
        if not m:
            return np.nan
        row = m.get(metric)
        if not row:
            return np.nan
        return row.get(col, np.nan)

    # ---- 집계표 구성 ----
    st.subheader("매니저별 진척 현황")
    if len(active_months) < len(MONTH_ORDER):
        st.caption(f"결산 데이터가 있는 월({', '.join(active_months)})만 표시하고 있어요. 나머지 월은 데이터가 들어오면 자동으로 나타나요.")

    col1, col2 = st.columns(2)
    with col1:
        manager_choice = st.selectbox("담당자 선택 (총합은 항상 표시돼요)", ["전체"] + managers, key="manager_kpi_manager_filter")
    with col2:
        month_choice = st.selectbox("표시할 월", ["전체"] + active_months, key="manager_kpi_month_filter")
    table_months = active_months if month_choice == "전체" else [month_choice]

    METRICS = ["매출 KPI", "GP KPI", "매출 결과", "GP 결과", "매출 달성률(%)", "GP 달성률(%)"]
    COLS = ["합계"] + table_months  # 화면 표시용 (비활성 월 숨김 + 선택한 월만)

    def build_group_rows(sales_kpi_row, gp_kpi_row, rev_row, gp_row):
        rows = {}
        rows["매출 KPI"] = sales_kpi_row
        rows["GP KPI"] = gp_kpi_row
        rows["매출 결과"] = rev_row
        rows["GP 결과"] = gp_row
        rows["매출 달성률(%)"] = {c: _safe_div(rev_row[c], sales_kpi_row[c]) for c in COLS}
        rows["GP 달성률(%)"] = {c: _safe_div(gp_row[c], gp_kpi_row[c]) for c in COLS}
        return rows

    groups = []

    # 부서 총합 (KPI는 백데이터의 '부서 총합' 행을 그대로 사용, 실적은 담당자 실적 합산)
    total_sales_kpi = {c: _kpi_value("부서 총합", "매출 KPI", c) for c in COLS}
    total_gp_kpi = {c: _kpi_value("부서 총합", "GP KPI", c) for c in COLS}
    total_rev = {c: (actuals.xs("매출 결과", level="지표")["합계"].sum() if c == "합계" else actuals.xs("매출 결과", level="지표")[c].sum()) for c in COLS}
    total_gp = {c: (actuals.xs("GP 결과", level="지표")["합계"].sum() if c == "합계" else actuals.xs("GP 결과", level="지표")[c].sum()) for c in COLS}
    groups.append(("부서 총합", build_group_rows(total_sales_kpi, total_gp_kpi, total_rev, total_gp)))

    for manager in managers:
        sales_kpi_row = {c: _kpi_value(manager, "매출 KPI", c) for c in COLS}
        gp_kpi_row = {c: _kpi_value(manager, "GP KPI", c) for c in COLS}
        rev_row = {c: actuals.loc[(manager, "매출 결과"), c] for c in COLS}
        gp_row = {c: actuals.loc[(manager, "GP 결과"), c] for c in COLS}
        groups.append((manager, build_group_rows(sales_kpi_row, gp_kpi_row, rev_row, gp_row)))

    if manager_choice != "전체":
        groups = [g for g in groups if g[0] in ("부서 총합", manager_choice)]

    # ---- 분기 헤더 구성 (선택된 월 기준) ----
    quarter_groups = []
    for q in ["1Q", "2Q", "3Q", "4Q"]:
        months_in_q = [m for m in table_months if QUARTER_MAP[m] == q]
        if months_in_q:
            quarter_groups.append((q, months_in_q))

    # ---- HTML 렌더링 ----
    html = """
    <style>
    .mgr-table-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    .mgr-table { border-collapse: collapse; width: 100%; font-size: 15px; }
    .mgr-table th, .mgr-table td {
        text-align: center !important;
        padding: 10px 12px;
        border: 1px solid #d9dce3;
        color: #1f2937 !important;
        white-space: nowrap;
    }
    .mgr-table thead th {
        background-color: #1f2a44 !important;
        color: #ffffff !important;
        font-weight: 700;
        font-size: 15px;
    }
    .mgr-table td.rowlabel {
        background-color: #f3f4f8 !important;
        font-weight: 600;
    }
    .mgr-table td.manager-cell {
        background-color: #e4e8f5 !important;
        font-weight: 700;
        font-size: 16px;
    }
    .mgr-table td.total-cell {
        background-color: #eef1fb !important;
        font-weight: 700;
    }
    .mgr-table tr.dept-total td {
        background-color: #fdf3e0 !important;
        font-weight: 700;
    }
    </style>
    <div class="mgr-table-wrap">
    <table class="mgr-table">
    <thead>
    <tr>
    <th rowspan="2">담당자</th><th rowspan="2">구분</th><th rowspan="2">합계</th>
    """
    for q, months_in_q in quarter_groups:
        html += f'<th colspan="{len(months_in_q)}">{q}</th>'
    html += "</tr><tr>"
    for q, months_in_q in quarter_groups:
        for month in months_in_q:
            html += f"<th>{month}</th>"
    html += "</tr></thead><tbody>"

    for gname, rows in groups:
        is_total = (gname == "부서 총합")
        for i, metric in enumerate(METRICS):
            row_class = ' class="dept-total"' if is_total else ""
            html += f"<tr{row_class}>"
            if i == 0:
                cell_cls = "total-cell" if is_total else "manager-cell"
                html += f'<td rowspan="{len(METRICS)}" class="{cell_cls}">{gname}</td>'
            html += f'<td class="rowlabel">{metric}</td>'
            for c in COLS:
                v = rows[metric][c]
                if "달성률" in metric:
                    text = _fmt_pct(v)
                else:
                    text = _fmt_money(v)
                cls = ' class="total-cell"' if c == "합계" else ""
                html += f"<td{cls}>{text}</td>"
            html += "</tr>"

    html += "</tbody></table></div>"

    st.markdown(html, unsafe_allow_html=True)

    # ---- 엑셀 다운로드 ----
    excel_bytes = _build_excel(groups, COLS, quarter_groups, METRICS)
    st.download_button(
        "📥 엑셀로 다운로드",
        data=excel_bytes,
        file_name="매니저별_진척현황.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _build_excel(groups, COLS, quarter_groups, METRICS) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "매니저별"

    # 헤더 (2행): 담당자/구분/합계는 세로 병합, 분기는 가로 병합
    r1, r2 = 1, 2
    xx.style_header(ws.cell(row=r1, column=1, value="담당자"))
    xx.style_header(ws.cell(row=r1, column=2, value="구분"))
    xx.style_header(ws.cell(row=r1, column=3, value="합계"))
    ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
    ws.merge_cells(start_row=r1, start_column=2, end_row=r2, end_column=2)
    ws.merge_cells(start_row=r1, start_column=3, end_row=r2, end_column=3)
    for r in (r1, r2):
        for c in (1, 2, 3):
            xx.style_header(ws.cell(row=r, column=c))
    ws.cell(row=r1, column=1, value="담당자")
    ws.cell(row=r1, column=2, value="구분")
    ws.cell(row=r1, column=3, value="합계")

    col = 4
    month_col_map = {}
    for q, months_in_q in quarter_groups:
        start_col = col
        for m in months_in_q:
            xx.style_header(ws.cell(row=r2, column=col, value=m))
            month_col_map[m] = col
            col += 1
        end_col = col - 1
        xx.style_header(ws.cell(row=r1, column=start_col, value=q))
        if end_col > start_col:
            ws.merge_cells(start_row=r1, start_column=start_col, end_row=r1, end_column=end_col)
            for c in range(start_col, end_col + 1):
                xx.style_header(ws.cell(row=r1, column=c))

    col_to_key = {3: "합계"}
    col_to_key.update({v: k for k, v in month_col_map.items()})
    last_col = col - 1

    row = 3
    for gname, rows in groups:
        is_total = (gname == "부서 총합")
        start_row = row
        for metric in METRICS:
            name_cell = ws.cell(row=row, column=1, value=gname if row == start_row else None)
            ws.cell(row=row, column=2, value=metric)
            for c in range(3, last_col + 1):
                key = col_to_key[c]
                v = rows[metric][key]
                is_pct = "달성률" in metric
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    val = None
                else:
                    val = float(v) if is_pct else round(v)
                fmt = xx.PCT_FMT if is_pct else xx.MONEY_FMT
                cell = ws.cell(row=row, column=c, value=val)
                if is_total:
                    xx.style_allrow(cell, number_format=fmt)
                elif c == 3:
                    xx.style_total(cell, number_format=fmt)
                else:
                    xx.style_plain(cell, number_format=fmt)

            if is_total:
                xx.style_allrow(ws.cell(row=row, column=2))
            else:
                xx.style_label(ws.cell(row=row, column=2))
            row += 1

        cell1 = ws.cell(row=start_row, column=1)
        if is_total:
            xx.style_allrow(cell1)
        else:
            xx.style_group(cell1)
        if len(METRICS) > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + len(METRICS) - 1, end_column=1)
            for r in range(start_row, start_row + len(METRICS)):
                c = ws.cell(row=r, column=1)
                if is_total:
                    xx.style_allrow(c)
                else:
                    xx.style_group(c)

    xx.autosize(ws)
    xx.freeze_header(ws, row=2, col=3)
    return xx.to_bytes(wb)
