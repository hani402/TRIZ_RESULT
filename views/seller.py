import streamlit as st
from openpyxl import Workbook

from data_utils import build_seller_view, build_seller_total, get_active_months, get_products, MONTH_ORDER
import excel_export as xx


def _fmt_money(v):
    if v is None:
        return ""
    if v == 0:
        return "-"
    return f"{int(round(v)):,}원"


def _fmt_pct(v):
    if v is None:
        return ""
    return f"{v * 100:,.2f}%"


def render(df):
    st.title("🙋 셀러별")

    if df is None:
        st.info("왼쪽 사이드바에서 ALL데이터 엑셀 파일을 업로드하면 결과가 표시됩니다.")
        return

    active_months = get_active_months(df)
    if len(active_months) < len(MONTH_ORDER):
        st.caption(f"결산 데이터가 있는 월({', '.join(active_months)})만 표시하고 있어요.")

    products = get_products(df)
    col1, col2 = st.columns(2)
    with col1:
        product_choice = st.selectbox("상품 선택 (선택 안 하면 전체 상품 합산)", ["전체"] + products)
    with col2:
        month_choice = st.selectbox("표시할 월", ["전체"] + active_months, key="seller_month_filter")
    table_months = active_months if month_choice == "전체" else [month_choice]

    rows = build_seller_view(df, active_months, product=product_choice)

    seller_names = [r["셀러명"] for r in rows]
    seller_choice = st.selectbox("셀러 선택 (총합은 항상 표시돼요)", ["전체"] + seller_names, key="seller_seller_filter")
    if seller_choice != "전체":
        rows = [r for r in rows if r["셀러명"] == seller_choice]

    total_row = build_seller_total(rows, active_months)
    st.caption(f"총 {len(rows):,}명의 셀러가 집계됐어요. (매출 높은 순 정렬, 맨 위는 전체 총합)")

    html = """
    <style>
    .sl-table-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    .sl-table { border-collapse: collapse; width: 100%; font-size: 15px; }
    .sl-table th, .sl-table td {
        text-align: center !important;
        padding: 9px 11px;
        border: 1px solid #d9dce3;
        color: #1f2937 !important;
        white-space: nowrap;
    }
    .sl-table thead th {
        background-color: #1f2a44 !important;
        color: #ffffff !important;
        font-weight: 600;
    }
    .sl-table td.info-cell {
        background-color: #f3f4f8 !important;
        font-weight: 600;
    }
    .sl-table td.total-cell {
        background-color: #eef1fb !important;
        font-weight: 700;
    }
    .sl-table tbody tr:nth-child(even) td:not(.info-cell):not(.total-cell) { background-color: #fafafc !important; }
    .sl-table tr.grand-total-row td {
        background-color: #fdf3e0 !important;
        font-weight: 700;
    }
    </style>
    <div class="sl-table-wrap">
    <table class="sl-table">
    <thead>
    <tr>
    <th rowspan="2">다이렉트/벤더</th><th rowspan="2">셀러명</th>
    <th colspan="6">총합</th>
    """
    for m in table_months:
        html += f'<th colspan="2">{m}</th>'
    html += "</tr><tr>"
    html += "<th>매출 비중</th><th>GP 비중</th><th>매출</th><th>평균 매출</th><th>트리즈 GP</th><th>평균 GP</th>"
    for m in table_months:
        html += "<th>매출</th><th>트리즈 GP</th>"
    html += "</tr></thead><tbody>"

    def _row_html(row, is_total=False):
        cls = ' class="grand-total-row"' if is_total else ""
        out = f"<tr{cls}>"
        out += f'<td class="info-cell">{row["다이렉트/벤더"]}</td>'
        out += f'<td class="info-cell">{row["셀러명"]}</td>'
        out += f'<td class="total-cell">{_fmt_pct(row["매출비중"])}</td>'
        out += f'<td class="total-cell">{_fmt_pct(row["GP비중"])}</td>'
        out += f'<td class="total-cell">{_fmt_money(row["매출"])}</td>'
        out += f'<td class="total-cell">{_fmt_money(row["평균매출"])}</td>'
        out += f'<td class="total-cell">{_fmt_money(row["GP"])}</td>'
        out += f'<td class="total-cell">{_fmt_money(row["평균GP"])}</td>'
        for m in table_months:
            out += f'<td>{_fmt_money(row.get(f"{m}_매출"))}</td>'
            out += f'<td>{_fmt_money(row.get(f"{m}_GP"))}</td>'
        out += "</tr>"
        return out

    html += _row_html(total_row, is_total=True)
    for row in rows:
        html += _row_html(row)

    html += "</tbody></table></div>"

    st.markdown(html, unsafe_allow_html=True)

    excel_bytes = _build_excel(total_row, rows, table_months)
    st.download_button(
        "📥 엑셀로 다운로드",
        data=excel_bytes,
        file_name="셀러별.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _build_excel(total_row, rows, table_months) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "셀러별"

    r1, r2 = 1, 2
    headers = ["다이렉트/벤더", "셀러명"]
    for c, text in enumerate(headers, start=1):
        ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
        xx.style_header(ws.cell(row=r1, column=c, value=text))
        xx.style_header(ws.cell(row=r2, column=c))

    total_headers = ["매출 비중", "GP 비중", "매출", "평균 매출", "트리즈 GP", "평균 GP"]
    start_col = 3
    xx.style_header(ws.cell(row=r1, column=start_col, value="총합"))
    ws.merge_cells(start_row=r1, start_column=start_col, end_row=r1, end_column=start_col + len(total_headers) - 1)
    for i, h in enumerate(total_headers):
        xx.style_header(ws.cell(row=r1, column=start_col + i))
        xx.style_header(ws.cell(row=r2, column=start_col + i, value=h))

    col = start_col + len(total_headers)
    for m in table_months:
        xx.style_header(ws.cell(row=r1, column=col, value=m))
        ws.merge_cells(start_row=r1, start_column=col, end_row=r1, end_column=col + 1)
        xx.style_header(ws.cell(row=r1, column=col + 1))
        xx.style_header(ws.cell(row=r2, column=col, value="매출"))
        xx.style_header(ws.cell(row=r2, column=col + 1, value="트리즈 GP"))
        col += 2

    def _write_row(row_idx, row, is_total):
        ws.cell(row=row_idx, column=1, value=row["다이렉트/벤더"])
        ws.cell(row=row_idx, column=2, value=row["셀러명"])
        vals = [row["매출비중"], row["GP비중"], row["매출"], row["평균매출"], row["GP"], row["평균GP"]]
        fmts = [xx.PCT_FMT, xx.PCT_FMT, xx.MONEY_FMT, xx.MONEY_FMT, xx.MONEY_FMT, xx.MONEY_FMT]
        for i, (v, fmt) in enumerate(zip(vals, fmts)):
            c = 3 + i
            val = None if v is None else (float(v) if "비중" in total_headers[i] else round(v))
            cell = ws.cell(row=row_idx, column=c, value=val)
            if is_total:
                xx.style_allrow(cell, number_format=fmt)
            else:
                xx.style_total(cell, number_format=fmt)
        c = 3 + len(total_headers)
        for m in table_months:
            for suffix in ("_매출", "_GP"):
                v = row.get(f"{m}{suffix}")
                cell = ws.cell(row=row_idx, column=c, value=(round(v) if v is not None else None))
                if is_total:
                    xx.style_allrow(cell, number_format=xx.MONEY_FMT)
                else:
                    xx.style_plain(cell, number_format=xx.MONEY_FMT)
                c += 1
        for c in (1, 2):
            cell = ws.cell(row=row_idx, column=c)
            if is_total:
                xx.style_allrow(cell)
            else:
                xx.style_label(cell)

    row_idx = 3
    _write_row(row_idx, total_row, is_total=True)
    row_idx += 1
    for row in rows:
        _write_row(row_idx, row, is_total=False)
        row_idx += 1

    xx.autosize(ws)
    xx.freeze_header(ws, row=2, col=2)
    return xx.to_bytes(wb)
