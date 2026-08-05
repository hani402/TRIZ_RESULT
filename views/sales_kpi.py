import streamlit as st
from plotly.subplots import make_subplots
from openpyxl import Workbook

from data_utils import build_sales_kpi, get_active_months, MONTH_ORDER
import excel_export as xx


def render(df):
    st.title("📊 영업 지표")

    if df is None:
        st.info("왼쪽 사이드바에서 ALL데이터 엑셀 파일을 업로드하면 결과가 표시됩니다.")
        return

    active_months = get_active_months(df)
    st.success(f"총 {len(df):,}건의 데이터를 불러왔어요. (데이터가 있는 월: {', '.join(active_months)})")
    if len(active_months) < len(MONTH_ORDER):
        st.caption("결산 데이터가 없는 월은 표/그래프에서 숨겼어요. 데이터가 들어오면 자동으로 나타나요.")

    kpi = build_sales_kpi(df)

    # ---- 표 ----
    st.subheader("월별 집계표")
    month_choice = st.selectbox("표시할 월", ["전체"] + active_months, key="sales_kpi_month_filter")
    table_months = active_months if month_choice == "전체" else [month_choice]

    display_cols = ["ALL"] + table_months
    display = kpi[display_cols].copy()
    for col in display.columns:
        display[col] = display.apply(
            lambda row: (
                "-" if int(row[col]) == 0
                else (f"{int(row[col]):,}" if row.name == "진행 건수" else f"{int(row[col]):,}원")
            ),
            axis=1,
        )

    table_html = """
    <style>
    .kpi-table-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    .kpi-table { border-collapse: collapse; width: 100%; font-size: 16px; }
    .kpi-table th, .kpi-table td {
        text-align: center !important;
        padding: 12px 14px;
        border: 1px solid #d9dce3;
        color: #1f2937 !important;
        white-space: nowrap;
    }
    .kpi-table thead th {
        background-color: #1f2a44 !important;
        color: #ffffff !important;
        font-weight: 600;
    }
    .kpi-table tbody th {
        background-color: #f3f4f8 !important;
        color: #1f2937 !important;
        font-weight: 600;
        text-align: center !important;
    }
    .kpi-table tbody tr:nth-child(even) td:not(.all-col) { background-color: #fafafc !important; }
    .kpi-table tbody tr:nth-child(odd) td:not(.all-col) { background-color: #ffffff !important; }
    .kpi-table td.all-col {
        background-color: #eef1fb !important;
        color: #1f2937 !important;
        font-weight: 700;
    }
    </style>
    <div class="kpi-table-wrap">
    <table class="kpi-table">
    <thead><tr><th>구분</th>"""
    for col in display.columns:
        table_html += f"<th>{col}</th>"
    table_html += "</tr></thead><tbody>"
    for idx, row in display.iterrows():
        table_html += f"<tr><th>{idx}</th>"
        for col in display.columns:
            cls = ' class="all-col"' if col == "ALL" else ""
            table_html += f"<td{cls}>{row[col]}</td>"
        table_html += "</tr>"
    table_html += "</tbody></table></div>"

    st.markdown(table_html, unsafe_allow_html=True)

    # ---- 차트 ----
    st.subheader("월별 추이")
    months = active_months

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_bar(
        x=months, y=[kpi.loc["매출 결과", m] for m in months],
        name="매출 결과", secondary_y=False,
    )
    fig.add_bar(
        x=months, y=[kpi.loc["GP 결과", m] for m in months],
        name="GP 결과", secondary_y=False,
    )
    fig.add_scatter(
        x=months, y=[kpi.loc["진행 건수", m] for m in months],
        name="진행 건수", mode="lines+markers", secondary_y=True,
        line=dict(width=3),
    )
    fig.update_layout(title="매출 · GP · 진행 건수", barmode="group", height=460)
    fig.update_yaxes(title_text="금액 (원)", secondary_y=False)
    fig.update_yaxes(title_text="진행 건수", secondary_y=True)
    st.plotly_chart(fig, use_container_width=True)

    # ---- 다운로드 ----
    st.subheader("엑셀로 다운로드")
    excel_bytes = _build_excel(kpi, table_months)
    st.download_button(
        "영업 지표.xlsx 다운로드",
        data=excel_bytes,
        file_name="영업_지표.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _build_excel(kpi, table_months) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "영업 지표"

    cols = ["ALL"] + table_months
    header_row = 1

    xx.style_header(ws.cell(row=header_row, column=1, value="구분"))
    for j, col in enumerate(cols, start=2):
        xx.style_header(ws.cell(row=header_row, column=j, value=col))

    metrics = ["진행 건수", "매출 결과", "GP 결과"]
    row_idx = {}
    for i, metric in enumerate(metrics, start=2):
        row_idx[metric] = i
        xx.style_label(ws.cell(row=i, column=1, value=metric))
        for j, col in enumerate(cols, start=2):
            v = kpi.loc[metric, col]
            fmt = xx.INT_FMT if metric == "진행 건수" else xx.MONEY_FMT
            cell = ws.cell(row=i, column=j, value=(int(v) if v is not None else None))
            if col == "ALL":
                xx.style_total(cell, number_format=fmt)
            else:
                xx.style_plain(cell, number_format=fmt)

    xx.autosize(ws)
    xx.freeze_header(ws, row=1, col=1)

    # 차트는 실제 월이 2개 이상 있을 때만 의미가 있어 추가
    if len(table_months) >= 1:
        first_col = 3  # ALL 다음 첫 월 컬럼
        last_col = 2 + len(table_months)
        xx.add_sales_kpi_chart(
            ws, header_row=header_row, first_data_col=first_col, last_data_col=last_col,
            rev_row=row_idx["매출 결과"], gp_row=row_idx["GP 결과"], count_row=row_idx["진행 건수"],
            anchor=f"A{len(metrics) + 4}",
        )

    return xx.to_bytes(wb)
