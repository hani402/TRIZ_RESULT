"""화면에 보이는 표(병합/서식)를 그대로 엑셀로 내보내기 위한 공통 헬퍼."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FILL = PatternFill("solid", fgColor="F3F4F8")
GROUP_FILL = PatternFill("solid", fgColor="E4E8F5")
TOTAL_FILL = PatternFill("solid", fgColor="EEF1FB")
ALLROW_FILL = PatternFill("solid", fgColor="FDF3E0")
PB_FILL = PatternFill("solid", fgColor="CFE3FF")
NB_FILL = PatternFill("solid", fgColor="CDF0DA")
PB_SUBTOTAL_FILL = PatternFill("solid", fgColor="A9CBFF")
NB_SUBTOTAL_FILL = PatternFill("solid", fgColor="9FE0B8")

THIN = Side(style="thin", color="D9DCE3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

MONEY_FMT = '#,##0"원";-#,##0"원";"-"'
PCT_FMT = "0.00%"
INT_FMT = '#,##0;-#,##0;"-"'


def _apply(cell, fill=None, font=None, number_format=None):
    cell.alignment = CENTER
    cell.border = BORDER
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if number_format is not None:
        cell.number_format = number_format


def style_header(cell):
    _apply(cell, fill=HEADER_FILL, font=HEADER_FONT)


def style_label(cell, bold=True):
    _apply(cell, fill=LABEL_FILL, font=Font(bold=bold))


def style_group(cell):
    _apply(cell, fill=GROUP_FILL, font=Font(bold=True))


def style_group_pb(cell):
    _apply(cell, fill=PB_FILL, font=Font(bold=True, color="1E40AF"))


def style_group_nb(cell):
    _apply(cell, fill=NB_FILL, font=Font(bold=True, color="166534"))


def style_subtotal_pb(cell, number_format=None):
    _apply(cell, fill=PB_SUBTOTAL_FILL, font=Font(bold=True), number_format=number_format)


def style_subtotal_nb(cell, number_format=None):
    _apply(cell, fill=NB_SUBTOTAL_FILL, font=Font(bold=True), number_format=number_format)


def style_total(cell, number_format=None):
    _apply(cell, fill=TOTAL_FILL, font=Font(bold=True), number_format=number_format)


def style_allrow(cell, number_format=None):
    _apply(cell, fill=ALLROW_FILL, font=Font(bold=True), number_format=number_format)


def style_plain(cell, number_format=None):
    _apply(cell, number_format=number_format)


def autosize(ws, min_width=9, max_width=24):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            col_letter = cell.column_letter
            widths[col_letter] = max(widths.get(col_letter, 0), len(str(v)))
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, w + 2))


def freeze_header(ws, row: int = 1, col: int = 1):
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = f"{get_column_letter(col + 1)}{row + 1}"


def to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def add_sales_kpi_chart(ws, header_row: int, first_data_col: int, last_data_col: int,
                         rev_row: int, gp_row: int, count_row: int, anchor: str):
    """매출/GP 막대 + 진행 건수 선그래프 콤보 차트를 추가. (ALL/라벨 열은 데이터에서 제외)"""
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.data_source import StrRef

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "매출 · GP · 진행 건수"
    bar.y_axis.title = "금액 (원)"
    bar.x_axis.title = "월"

    cats = Reference(ws, min_col=first_data_col, max_col=last_data_col, min_row=header_row, max_row=header_row)

    rev_ref = Reference(ws, min_col=first_data_col, max_col=last_data_col, min_row=rev_row, max_row=rev_row)
    bar.add_data(rev_ref, titles_from_data=False)
    bar.series[-1].tx = SeriesLabel(v="매출 결과")

    gp_ref = Reference(ws, min_col=first_data_col, max_col=last_data_col, min_row=gp_row, max_row=gp_row)
    bar.add_data(gp_ref, titles_from_data=False)
    bar.series[-1].tx = SeriesLabel(v="GP 결과")

    bar.set_categories(cats)

    line = LineChart()
    cnt_ref = Reference(ws, min_col=first_data_col, max_col=last_data_col, min_row=count_row, max_row=count_row)
    line.add_data(cnt_ref, titles_from_data=False)
    line.series[-1].tx = SeriesLabel(v="진행 건수")
    line.y_axis.axId = 200
    line.y_axis.title = "진행 건수"
    line.y_axis.crosses = "max"

    bar += line
    bar.height = 9
    bar.width = 22
    ws.add_chart(bar, anchor)
