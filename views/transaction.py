import streamlit as st
from openpyxl import Workbook

from data_utils import build_transaction_view, get_active_months, QUARTER_MAP, MONTH_ORDER
import excel_export as xx


def _fmt_money(v):
    if v is None:
        return ""
    return f"{int(round(v)):,}원"


def _fmt_count(v):
    if v is None:
        return ""
    return f"{int(v):,}"


def render(df):
    st.title("📑 거래별")

    if df is None:
        st.info("왼쪽 사이드바에서 ALL데이터 엑셀 파일을 업로드하면 결과가 표시됩니다.")
        return

    active_months = get_active_months(df)
    if len(active_months) < len(MONTH_ORDER):
        st.caption(f"결산 데이터가 있는 월({', '.join(active_months)})만 표시하고 있어요.")

    blocks = build_transaction_view(df, active_months)

    def _block_label(b):
        return f"{b['category']} - {b['sub']}" if b["sub"] else b["category"]

    options = [_block_label(b) for b in blocks if b["category"] != "ALL"]

    col1, col2 = st.columns(2)
    with col1:
        category_choice = st.selectbox("구분 선택 (ALL은 항상 표시돼요)", ["전체"] + options, key="transaction_category_filter")
    with col2:
        month_choice = st.selectbox("표시할 월", ["전체"] + active_months, key="transaction_month_filter")
    table_months = active_months if month_choice == "전체" else [month_choice]

    if category_choice != "전체":
        blocks = [b for b in blocks if b["category"] == "ALL" or _block_label(b) == category_choice]

    COLS = ["합계"] + table_months
    quarter_groups = []
    for q in ["1Q", "2Q", "3Q", "4Q"]:
        months_in_q = [m for m in table_months if QUARTER_MAP[m] == q]
        if months_in_q:
            quarter_groups.append((q, months_in_q))

    # ---- 다이렉트/벤더(대분류) rowspan 계산 ----
    category_rowcount = {}
    for b in blocks:
        n_metric_rows = len(b["metrics"])
        category_rowcount[b["category"]] = category_rowcount.get(b["category"], 0) + n_metric_rows

    html = """
    <style>
    .tx-table-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    .tx-table { border-collapse: collapse; width: 100%; font-size: 15px; }
    .tx-table th, .tx-table td {
        text-align: center !important;
        padding: 9px 11px;
        border: 1px solid #d9dce3;
        color: #1f2937 !important;
        white-space: nowrap;
    }
    .tx-table thead th {
        background-color: #1f2a44 !important;
        color: #ffffff !important;
        font-weight: 600;
    }
    .tx-table td.cat-cell {
        background-color: #e4e8f5 !important;
        font-weight: 700;
    }
    .tx-table td.sub-cell {
        background-color: #f3f4f8 !important;
        font-weight: 600;
    }
    .tx-table td.metric-cell {
        background-color: #fafafc !important;
    }
    .tx-table td.total-cell {
        background-color: #eef1fb !important;
        font-weight: 700;
    }
    .tx-table tr.all-row td {
        background-color: #fdf3e0 !important;
        font-weight: 700;
    }
    .tx-table td.vendor-sub {
        color: #6b7280 !important;
        font-weight: 500;
    }
    </style>
    <div class="tx-table-wrap">
    <table class="tx-table">
    <thead>
    <tr>
    <th rowspan="2">다이렉트/벤더</th><th rowspan="2">세부</th><th rowspan="2">인원수</th>
    <th rowspan="2">지표</th><th rowspan="2">합계</th>
    """
    for q, months_in_q in quarter_groups:
        html += f'<th colspan="{len(months_in_q)}">{q}</th>'
    html += "</tr><tr>"
    for q, months_in_q in quarter_groups:
        for month in months_in_q:
            html += f"<th>{month}</th>"
    html += "</tr></thead><tbody>"

    last_emitted_category = None
    for b in blocks:
        is_all = b["category"] == "ALL"
        metrics = b["metrics"]
        metric_names = list(metrics.keys())
        n_rows = len(metric_names)

        for i, metric in enumerate(metric_names):
            row_class = ' class="all-row"' if is_all else ""
            html += f"<tr{row_class}>"

            if b["category"] != last_emitted_category:
                html += f'<td rowspan="{category_rowcount[b["category"]]}" class="cat-cell">{b["category"]}</td>'
                last_emitted_category = b["category"]

            if i == 0:
                sub_text = b["sub"] if b["sub"] else "-"
                sub_cls = "sub-cell vendor-sub" if (b["category"] == "벤더사" and b["sub"] and "전체" not in b["sub"]) else "sub-cell"
                html += f'<td rowspan="{n_rows}" class="{sub_cls}">{sub_text}</td>'
                html += f'<td rowspan="{n_rows}" class="sub-cell">{b["headcount"]}명</td>'

            html += f'<td class="metric-cell">{metric}</td>'

            row_vals = metrics[metric]
            for c in COLS:
                v = row_vals.get(c)
                text = _fmt_count(v) if metric == "진행 횟수" else _fmt_money(v)
                cls = ' class="total-cell"' if c == "합계" else ""
                html += f"<td{cls}>{text}</td>"
            html += "</tr>"

    html += "</tbody></table></div>"

    st.markdown(html, unsafe_allow_html=True)

    excel_bytes = _build_excel(blocks, COLS, quarter_groups, category_rowcount)
    st.download_button(
        "📥 엑셀로 다운로드",
        data=excel_bytes,
        file_name="거래별.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _build_excel(blocks, COLS, quarter_groups, category_rowcount) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "거래별"

    r1, r2 = 1, 2
    headers = ["다이렉트/벤더", "세부", "인원수", "지표", "합계"]
    for c, text in enumerate(headers, start=1):
        ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
        ws.cell(row=r1, column=c, value=text)
        for r in (r1, r2):
            xx.style_header(ws.cell(row=r, column=c))

    col = 6
    month_col_map = {}
    for q, months_in_q in quarter_groups:
        start_col = col
        for m in months_in_q:
            xx.style_header(ws.cell(row=r2, column=col, value=m))
            month_col_map[m] = col
            col += 1
        end_col = col - 1
        xx.style_header(ws.cell(row=r1, column=start_col, value=q))
        if end_col > start_col:
            ws.merge_cells(start_row=r1, start_column=start_col, end_row=r1, end_column=end_col)
            for c in range(start_col, end_col + 1):
                xx.style_header(ws.cell(row=r1, column=c))

    col_to_key = {5: "합계"}
    col_to_key.update({v: k for k, v in month_col_map.items()})
    last_col = col - 1

    row = 3
    last_emitted_category = None
    category_start_row = {}
    for b in blocks:
        is_all = b["category"] == "ALL"
        metrics = b["metrics"]
        metric_names = list(metrics.keys())
        n_rows = len(metric_names)

        if b["category"] != last_emitted_category:
            category_start_row[b["category"]] = row
            last_emitted_category = b["category"]

        for i, metric in enumerate(metric_names):
            if i == 0:
                sub_text = b["sub"] if b["sub"] else "-"
                ws.cell(row=row, column=2, value=sub_text)
                ws.cell(row=row, column=3, value=f'{b["headcount"]}명')
                if n_rows > 1:
                    ws.merge_cells(start_row=row, start_column=2, end_row=row + n_rows - 1, end_column=2)
                    ws.merge_cells(start_row=row, start_column=3, end_row=row + n_rows - 1, end_column=3)

            ws.cell(row=row, column=4, value=metric)

            row_vals = metrics[metric]
            for c in range(5, last_col + 1):
                key = col_to_key[c]
                v = row_vals.get(key)
                fmt = xx.INT_FMT if metric == "진행 횟수" else xx.MONEY_FMT
                val = int(v) if (metric == "진행 횟수" and v is not None) else (round(v) if v is not None else None)
                cell = ws.cell(row=row, column=c, value=val)
                if is_all:
                    xx.style_allrow(cell, number_format=fmt)
                elif c == 5:
                    xx.style_total(cell, number_format=fmt)
                else:
                    xx.style_plain(cell, number_format=fmt)

            for c in (2, 3, 4):
                cell = ws.cell(row=row, column=c)
                if is_all:
                    xx.style_allrow(cell)
                else:
                    xx.style_label(cell)
            row += 1


    # 카테고리(다이렉트/벤더) 세로 병합
    for category, total_rows in category_rowcount.items():
        start = category_start_row.get(category)
        if start is None:
            continue
        end = start + total_rows - 1
        ws.cell(row=start, column=1, value=category)
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        is_all = (category == "ALL")
        for r in range(start, end + 1):
            cell = ws.cell(row=r, column=1)
            if is_all:
                xx.style_allrow(cell)
            else:
                xx.style_group(cell)

    xx.autosize(ws)
    xx.freeze_header(ws, row=2, col=5)
    return xx.to_bytes(wb)
