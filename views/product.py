import streamlit as st
from openpyxl import Workbook

from data_utils import build_product_view, build_category_products, get_active_months, QUARTER_MAP, MONTH_ORDER
import excel_export as xx


def _fmt_money(v):
    if v is None:
        return ""
    return f"{int(round(v)):,}원"


def render(df):
    st.title("🏷️ 상품별")

    if df is None:
        st.info("왼쪽 사이드바에서 ALL데이터 엑셀 파일을 업로드하면 결과가 표시됩니다.")
        return

    active_months = get_active_months(df)
    if len(active_months) < len(MONTH_ORDER):
        st.caption(f"결산 데이터가 있는 월({', '.join(active_months)})만 표시하고 있어요.")

    blocks = build_product_view(df, active_months)

    def _block_label(b):
        return f"{b['group']} - {b['sub']}" if b["sub"] else b["group"]

    label_to_block = {_block_label(b): b for b in blocks if b["group"] != "ALL"}
    options = list(label_to_block.keys())

    col1, col2 = st.columns(2)
    with col1:
        group_choice = st.selectbox("구분 선택 (ALL은 항상 표시돼요)", ["전체"] + options, key="product_group_filter")
    with col2:
        month_choice = st.selectbox("표시할 월", ["전체"] + active_months, key="product_month_filter")
    table_months = active_months if month_choice == "전체" else [month_choice]

    if group_choice != "전체":
        blocks = [b for b in blocks if b["group"] == "ALL" or _block_label(b) == group_choice]

    COLS = ["합계"] + table_months
    quarter_groups = []
    for q in ["1Q", "2Q", "3Q", "4Q"]:
        months_in_q = [m for m in table_months if QUARTER_MAP[m] == q]
        if months_in_q:
            quarter_groups.append((q, months_in_q))

    group_rowcount = {}
    for b in blocks:
        group_rowcount[b["group"]] = group_rowcount.get(b["group"], 0) + len(b["metrics"])

    html = """
    <style>
    .pd-table-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    .pd-table { border-collapse: collapse; width: 100%; font-size: 15px; }
    .pd-table th, .pd-table td {
        text-align: center !important;
        padding: 9px 11px;
        border: 1px solid #d9dce3;
        color: #1f2937 !important;
        white-space: nowrap;
    }
    .pd-table thead th {
        background-color: #1f2a44 !important;
        color: #ffffff !important;
        font-weight: 600;
    }
    .pd-table td.group-cell {
        background-color: #e4e8f5 !important;
        font-weight: 700;
    }
    .pd-table td.group-cell-pb {
        background-color: #cfe3ff !important;
        color: #1e40af !important;
        font-weight: 700;
    }
    .pd-table td.group-cell-nb {
        background-color: #cdf0da !important;
        color: #166534 !important;
        font-weight: 700;
    }
    .pd-table td.sub-cell {
        background-color: #ffffff !important;
        font-weight: 600;
    }
    .pd-table td.metric-cell {
        background-color: #ffffff !important;
    }
    .pd-table td.total-cell {
        background-color: #eef1fb !important;
        font-weight: 700;
    }
    .pd-table tr.all-row td {
        background-color: #fdf3e0 !important;
        font-weight: 700;
    }
    .pd-table tr.subtotal-row-pb td {
        background-color: #a9cbff !important;
        font-weight: 700;
    }
    .pd-table tr.subtotal-row-nb td {
        background-color: #9fe0b8 !important;
        font-weight: 700;
    }
    </style>
    <div class="pd-table-wrap">
    <table class="pd-table">
    <thead>
    <tr>
    <th rowspan="2">구분</th><th rowspan="2">카테고리</th><th rowspan="2">지표</th><th rowspan="2">합계</th>
    """
    for q, months_in_q in quarter_groups:
        html += f'<th colspan="{len(months_in_q)}">{q}</th>'
    html += "</tr><tr>"
    for q, months_in_q in quarter_groups:
        for month in months_in_q:
            html += f"<th>{month}</th>"
    html += "</tr></thead><tbody>"

    SUBTOTAL_ROW_CLASS = {"PB": "subtotal-row-pb", "NB": "subtotal-row-nb"}

    last_emitted_group = None
    for b in blocks:
        is_all = b["group"] == "ALL"
        is_subtotal = b["sub"] == "소계"
        metrics = b["metrics"]
        metric_names = list(metrics.keys())
        n_rows = len(metric_names)

        for i, metric in enumerate(metric_names):
            row_class = ""
            if is_all:
                row_class = ' class="all-row"'
            elif is_subtotal:
                row_class = f' class="{SUBTOTAL_ROW_CLASS.get(b["group"], "subtotal-row")}"'
            html += f"<tr{row_class}>"

            if b["group"] != last_emitted_group:
                html += f'<td rowspan="{group_rowcount[b["group"]]}" class="group-cell">{b["group"]}</td>'
                last_emitted_group = b["group"]

            if i == 0:
                sub_text = b["sub"] if b["sub"] else "-"
                html += f'<td rowspan="{n_rows}" class="sub-cell">{sub_text}</td>'

            html += f'<td class="metric-cell">{metric}</td>'

            row_vals = metrics[metric]
            for c in COLS:
                v = row_vals.get(c)
                cls = ' class="total-cell"' if c == "합계" else ""
                html += f"<td{cls}>{_fmt_money(v)}</td>"
            html += "</tr>"

    html += "</tbody></table></div>"

    st.markdown(html, unsafe_allow_html=True)

    excel_bytes = _build_excel(blocks, COLS, quarter_groups, group_rowcount)
    st.download_button(
        "📥 엑셀로 다운로드",
        data=excel_bytes,
        file_name="상품별.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ---- 선택한 구분(카테고리) 안의 상품별 상세 ----
    if group_choice != "전체":
        selected_block = label_to_block[group_choice]
        product_rows = build_category_products(df, selected_block["group"], selected_block["sub"], table_months)

        st.subheader(f"'{group_choice}' 안의 상품별 매출/GP")
        if not product_rows:
            st.caption("해당 구분에 상품 데이터가 없어요.")
        else:
            st.caption(f"총 {len(product_rows):,}개 상품 (매출 높은 순 정렬)")
            detail_html = """
            <style>
            .pdd-table-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }
            .pdd-table { border-collapse: collapse; width: 100%; font-size: 15px; }
            .pdd-table th, .pdd-table td {
                text-align: center !important;
                padding: 9px 11px;
                border: 1px solid #d9dce3;
                color: #1f2937 !important;
                white-space: nowrap;
            }
            .pdd-table thead th {
                background-color: #1f2a44 !important;
                color: #ffffff !important;
                font-weight: 600;
            }
            .pdd-table td.name-cell {
                text-align: left !important;
                font-weight: 600;
            }
            .pdd-table td.total-cell {
                background-color: #eef1fb !important;
                font-weight: 700;
            }
            </style>
            <div class="pdd-table-wrap">
            <table class="pdd-table">
            <thead><tr><th>진행상품</th><th>매출</th><th>GP</th>
            """
            for m in table_months:
                detail_html += f"<th>{m} 매출</th><th>{m} GP</th>"
            detail_html += "</tr></thead><tbody>"
            for r in product_rows:
                detail_html += f'<tr><td class="name-cell">{r["진행상품"]}</td>'
                detail_html += f'<td class="total-cell">{_fmt_money(r["매출"])}</td>'
                detail_html += f'<td class="total-cell">{_fmt_money(r["GP"])}</td>'
                for m in table_months:
                    detail_html += f'<td>{_fmt_money(r.get(f"{m}_매출"))}</td>'
                    detail_html += f'<td>{_fmt_money(r.get(f"{m}_GP"))}</td>'
                detail_html += "</tr>"
            detail_html += "</tbody></table></div>"
            st.markdown(detail_html, unsafe_allow_html=True)


def _build_excel(blocks, COLS, quarter_groups, group_rowcount) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "상품별"

    r1, r2 = 1, 2
    headers = ["구분", "카테고리", "지표", "합계"]
    for c, text in enumerate(headers, start=1):
        ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
        xx.style_header(ws.cell(row=r1, column=c, value=text))
        xx.style_header(ws.cell(row=r2, column=c))

    col = 5
    month_col_map = {}
    for q, months_in_q in quarter_groups:
        start_col = col
        for m in months_in_q:
            xx.style_header(ws.cell(row=r2, column=col, value=m))
            month_col_map[m] = col
            col += 1
        end_col = col - 1
        xx.style_header(ws.cell(row=r1, column=start_col, value=q))
        if end_col > start_col:
            ws.merge_cells(start_row=r1, start_column=start_col, end_row=r1, end_column=end_col)
            for c in range(start_col, end_col + 1):
                xx.style_header(ws.cell(row=r1, column=c))

    col_to_key = {4: "합계"}
    col_to_key.update({v: k for k, v in month_col_map.items()})
    last_col = col - 1

    row = 3
    last_emitted_group = None
    group_start_row = {}
    for b in blocks:
        is_all = b["group"] == "ALL"
        is_subtotal = b["sub"] == "소계"
        metrics = b["metrics"]
        metric_names = list(metrics.keys())
        n_rows = len(metric_names)

        if b["group"] != last_emitted_group:
            group_start_row[b["group"]] = row
            last_emitted_group = b["group"]

        for i, metric in enumerate(metric_names):
            if i == 0:
                sub_text = b["sub"] if b["sub"] else "-"
                ws.cell(row=row, column=2, value=sub_text)
                if n_rows > 1:
                    ws.merge_cells(start_row=row, start_column=2, end_row=row + n_rows - 1, end_column=2)

            ws.cell(row=row, column=3, value=metric)

            row_vals = metrics[metric]
            for c in range(4, last_col + 1):
                key = col_to_key[c]
                v = row_vals.get(key)
                val = round(v) if v is not None else None
                cell = ws.cell(row=row, column=c, value=val)
                if is_all:
                    xx.style_allrow(cell, number_format=xx.MONEY_FMT)
                elif is_subtotal and b["group"] == "PB":
                    xx.style_subtotal_pb(cell, number_format=xx.MONEY_FMT)
                elif is_subtotal and b["group"] == "NB":
                    xx.style_subtotal_nb(cell, number_format=xx.MONEY_FMT)
                elif c == 4:
                    xx.style_total(cell, number_format=xx.MONEY_FMT)
                else:
                    xx.style_plain(cell, number_format=xx.MONEY_FMT)

            for c in (2, 3):
                cell = ws.cell(row=row, column=c)
                if is_all:
                    xx.style_allrow(cell)
                elif is_subtotal and b["group"] == "PB":
                    xx.style_subtotal_pb(cell)
                elif is_subtotal and b["group"] == "NB":
                    xx.style_subtotal_nb(cell)
                elif is_subtotal:
                    xx.style_total(cell)
                else:
                    xx.style_plain(cell)
            row += 1

    for group, total_rows in group_rowcount.items():
        start = group_start_row.get(group)
        if start is None:
            continue
        end = start + total_rows - 1
        ws.cell(row=start, column=1, value=group)
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        is_all = (group == "ALL")
        for r in range(start, end + 1):
            cell = ws.cell(row=r, column=1)
            if is_all:
                xx.style_allrow(cell)
            else:
                xx.style_group(cell)

    xx.autosize(ws)
    xx.freeze_header(ws, row=2, col=4)
    return xx.to_bytes(wb)
